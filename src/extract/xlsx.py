from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook


def extract(data: bytes) -> str:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        out.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                out.append(" | ".join(cells))
    return "\n".join(out)
